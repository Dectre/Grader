import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

class DataManager:
    def __init__(self):
        self.students_df = pd.read_csv(os.path.join("Data", "students.csv"), encoding='utf-8-sig')
        self.students_df['نام کاربری'] = self.students_df['نام کاربری'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        self.students_df = self.students_df.sort_values(by='نام خانوادگی')

        self.rubric_df = pd.read_csv(os.path.join("Data", "rubric.csv"), encoding='utf-8-sig')
        self.questions = self.rubric_df.iloc[:, 0].astype(str).tolist()
        self.max_grades = self.rubric_df.iloc[:, 1].tolist()

        self.output_dir = "output"
        self.excel_path = os.path.join(self.output_dir, "grades.xlsx")
        self.csv_path = os.path.join(self.output_dir, "grades.csv")
        self.pdf_dir = "PDFs"

        self.setup_dataframes()
        self.students = self.get_matched_students()

    def setup_dataframes(self):
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        cols = ['First Name', 'Last Name', 'Student ID'] + self.questions + ['Total Score', 'Description', '_Submitted']

        if os.path.exists(self.csv_path):
            self.grades_df = pd.read_csv(self.csv_path, encoding='utf-8-sig')
            self.grades_df['Student ID'] = self.grades_df['Student ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            if '_Submitted' not in self.grades_df.columns:
                self.grades_df['_Submitted'] = False
        elif os.path.exists(self.excel_path):
            df_temp = pd.read_excel(self.excel_path)
            
            c_names = list(df_temp.columns)
            if len(c_names) >= 3:
                c_names[0] = 'First Name'
                c_names[1] = 'Last Name'
                c_names[2] = 'Student ID'
                df_temp.columns = c_names
            
            df_temp = df_temp[~df_temp['First Name'].isin(['Mean', 'Median', 'Max', 'Min', 'First Name'])]
            df_temp = df_temp.dropna(subset=['Student ID'])
            
            self.grades_df = df_temp
            self.grades_df['Student ID'] = self.grades_df['Student ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            self.grades_df['_Submitted'] = False
        else:
            self.grades_df = pd.DataFrame(columns=cols)

        if 'Description' not in self.grades_df.columns:
            self.grades_df['Description'] = ""
        self.grades_df['Description'] = self.grades_df['Description'].astype('object')

        existing_ids = self.grades_df['Student ID'].tolist()
        new_rows = []
        for _, row in self.students_df.iterrows():
            sid = str(row['نام کاربری'])
            if sid != 'nan' and bool(sid) and sid not in existing_ids:
                new_rows.append({
                    'First Name': row['نام'],
                    'Last Name': row['نام خانوادگی'],
                    'Student ID': sid,
                    'Description': "",
                    '_Submitted': False
                })
        if new_rows:
            self.grades_df = pd.concat([self.grades_df, pd.DataFrame(new_rows)], ignore_index=True)

        self.grades_df = self.grades_df.sort_values(by='Last Name')
        self.export_files()

    def export_files(self):
        self.grades_df.to_csv(self.csv_path, index=False, encoding='utf-8-sig')
        export_df = self.grades_df.drop(columns=['_Submitted'], errors='ignore')
        export_df.to_excel(self.excel_path, index=False)
        self.apply_excel_formatting()

    def apply_excel_formatting(self):
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            ws.insert_rows(2, 5)
            
            vazir_font = Font(name='Vazirmatn')
            vazir_bold = Font(name='Vazirmatn', bold=True)
            
            center_align = Alignment(horizontal='center', vertical='center')
            desc_align = Alignment(horizontal='right', vertical='top', wrap_text=True)

            thick = Side(border_style="thick", color="000000")
            thin = Side(border_style="thin", color="000000")

            max_r = ws.max_row
            max_c = ws.max_column

            desc_col_idx = None
            total_col_idx = None
            q_col_indices = []
            
            ws.cell(row=1, column=1).value = ""
            ws.cell(row=1, column=2).value = ""
            ws.cell(row=1, column=3).value = ""
            
            ws.cell(row=6, column=1).value = "First Name"
            ws.cell(row=6, column=2).value = "Last Name"
            ws.cell(row=6, column=3).value = "Student ID"
            
            for c in range(1, max_c + 1):
                val = ws.cell(row=1, column=c).value
                if val == 'Description':
                    desc_col_idx = c
                elif val == 'Total Score':
                    total_col_idx = c
                elif val in self.questions:
                    q_col_indices.append(c)

            for c in range(1, max_c + 1):
                col_letter = ws.cell(row=1, column=c).column_letter
                if c == desc_col_idx:
                    ws.column_dimensions[col_letter].width = 50
                elif c in [1, 2, 3]:
                    ws.column_dimensions[col_letter].width = 18
                else:
                    ws.column_dimensions[col_letter].width = 12

            stat_labels = ["Mean", "Median", "Max", "Min"]
            for i in range(4):
                r = i + 2
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
                cell = ws.cell(row=r, column=1)
                cell.value = stat_labels[i]

            stat_cols = q_col_indices + ([total_col_idx] if total_col_idx else [])
            for c in stat_cols:
                col_name = ws.cell(row=1, column=c).value
                scores = pd.to_numeric(self.grades_df[col_name], errors='coerce').dropna()
                mean_val = round(scores.mean(), 2) if not scores.empty else 0
                median_val = round(scores.median(), 2) if not scores.empty else 0
                max_val = round(scores.max(), 2) if not scores.empty else 0
                min_val = round(scores.min(), 2) if not scores.empty else 0
                
                vals = [mean_val, median_val, max_val, min_val]
                for i in range(4):
                    r = i + 2
                    ws.cell(row=r, column=c).value = vals[i]

            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    
                    if r <= 6:
                        cell.font = vazir_bold
                    else:
                        cell.font = vazir_font
                    
                    if c == desc_col_idx and r > 6:
                        cell.alignment = desc_align
                    else:
                        cell.alignment = center_align

                    b_top = thin; b_bot = thin; b_left = thin; b_right = thin
                    
                    if r == 1:
                        b_top = thick
                        b_bot = thick
                    if r == 2: b_top = thick
                    if r == 6:
                        b_top = thick
                        b_bot = thick
                    if r == 7: b_top = thick
                    if r == max_r: b_bot = thick
                    
                    if c == 1: b_left = thick
                    if c == 3: b_right = thick
                    if c == 4: b_left = thick
                    if c == max_c: b_right = thick
                    
                    cell.border = Border(top=b_top, bottom=b_bot, left=b_left, right=b_right)

            ws.freeze_panes = 'D2'

            ws.sheet_view.rightToLeft = False
            wb.save(self.excel_path)
        except Exception:
            pass

    def get_matched_students(self):
        students_list = []
        for _, row in self.students_df.iterrows():
            student_id = row['نام کاربری']
            if student_id == 'nan' or not student_id:
                continue
            
            first_name = str(row['نام']).strip()
            last_name = str(row['نام خانوادگی']).strip()
            
            pdf_path = self.find_pdf(first_name, last_name)
            students_list.append({
                'name': first_name,
                'surname': last_name,
                'id': student_id,
                'pdf': pdf_path
                })
        return students_list

    def find_pdf(self, first_name, last_name):
        if not os.path.exists(self.pdf_dir):
            return None
            
        first = first_name.replace(" ", "").replace("‌", "")
        last = last_name.replace(" ", "").replace("‌", "")
            
        for filename in os.listdir(self.pdf_dir):
            if not filename.lower().endswith('.pdf'):
                continue
                
            clean_filename = filename.replace(" ", "").replace("‌", "").replace("_", "")
                
            if first in clean_filename and last in clean_filename:
                return os.path.join(self.pdf_dir, filename)
                
        return None

    def get_saved_data(self, student_id):
        idx = self.grades_df.index[self.grades_df['Student ID'] == str(student_id)].tolist()
        if idx:
            row = self.grades_df.iloc[idx[0]]
            grades = {}
            for q in self.questions:
                val = row.get(q, 0.0)
                grades[q] = 0.0 if pd.isna(val) else float(val)
            desc = row.get('Description', '')
            desc = "" if pd.isna(desc) else str(desc)
            return grades, desc
        return {q: 0.0 for q in self.questions}, ""
        
    def is_submitted(self, student_id):
        idx = self.grades_df.index[self.grades_df['Student ID'] == str(student_id)].tolist()
        if idx:
            val = self.grades_df.iloc[idx[0]].get('_Submitted', False)
            if pd.isna(val):
                return False
            return bool(val)
        return False

    def save_grade(self, student_id, grades_dict, total, comments):
        idx = self.grades_df.index[self.grades_df['Student ID'] == str(student_id)].tolist()
        if idx:
            row_idx = idx[0]
            for q in self.questions:
                self.grades_df.at[row_idx, q] = grades_dict.get(q, 0)
            self.grades_df.at[row_idx, 'Total Score'] = total
            self.grades_df.at[row_idx, 'Description'] = str(comments)
            self.grades_df.at[row_idx, '_Submitted'] = True

            self.export_files()